import openpyxl
import os

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active

# Add header
headers = ['name', 'description', 'price', 'sku', 'category', 'stock']
ws.append(headers)

# Add data - ensure stock is INTEGER type
data = [
    ['LED Desk Lamp', 'Bright LED lamp', 89.99, 'LAMP-001', 'Office', 45],
    ['Wireless Mouse', 'Ergonomic mouse', 29.99, 'MOUSE-001', 'Electronics', 120],
    ['Mechanical Keyboard', 'RGB keyboard', 149.99, 'KEY-001', 'Electronics', 60],
    ['USB-C Hub', 'USB hub HDMI', 49.99, 'HUB-001', 'Accessories', 85],
    ['Monitor Stand', 'Adjustable stand', 39.99, 'STAND-001', 'Office', 40],
]

for row in data:
    # Convert stock to int to ensure proper type
    row_copy = row.copy()
    row_copy[5] = int(row_copy[5])  # Stock column (index 5)
    ws.append(row_copy)

# Set column types explicitly
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[2].number_format = '0.00'  # Price as decimal
    row[5].number_format = '0'     # Stock as integer

# Save to project root
os.chdir('C:\\Users\\Administrator\\Platform Sales & Procurement')
wb.save('sample_products.xlsx')
print('✅ Created sample_products.xlsx with valid integer stock values')
